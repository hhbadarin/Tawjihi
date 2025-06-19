from openpyxl import load_workbook, Workbook
from openpyxl.styles import PatternFill, Font, Border, Alignment
import os
import shutil

# Load original workbook
input_path = os.path.expanduser('~/Desktop/توزيع الغياب حسب القاعة محدث.xlsx')
original_wb = load_workbook(input_path)

# Create output folder
output_folder = os.path.expanduser('~/Desktop/split_sheets')
os.makedirs(output_folder, exist_ok=True)

for sheet in original_wb.worksheets:
    # Create a new workbook
    new_wb = Workbook()
    new_ws = new_wb.active
    new_ws.title = sheet.title

    # Copy cell values and styles
    for row in sheet.iter_rows():
        for cell in row:
            new_cell = new_ws.cell(row=cell.row, column=cell.column, value=cell.value)

            # Copy styles
            if cell.has_style:
                new_cell.font = cell.font.copy()
                new_cell.border = cell.border.copy()
                new_cell.fill = cell.fill.copy()
                new_cell.number_format = cell.number_format
                new_cell.protection = cell.protection.copy()
                new_cell.alignment = cell.alignment.copy()

    # Copy column widths
    for col_letter, dim in sheet.column_dimensions.items():
        new_ws.column_dimensions[col_letter].width = dim.width

    # Copy row heights
    for row_idx, dim in sheet.row_dimensions.items():
        new_ws.row_dimensions[row_idx].height = dim.height

    # Copy page setup (margins, orientation, etc.)
    new_ws.page_margins = sheet.page_margins
    new_ws.page_setup = sheet.page_setup
    new_ws.sheet_properties.pageSetUpPr = sheet.sheet_properties.pageSetUpPr
    new_ws.print_title_rows = sheet.print_title_rows

    # Copy header/footer
    new_ws.oddHeader = sheet.oddHeader
    new_ws.oddFooter = sheet.oddFooter

    # Save the new file
    safe_title = "".join(c if c.isalnum() or c in " _-" else "_" for c in sheet.title)
    out_path = os.path.join(output_folder, f"{safe_title}.xlsx")
    new_wb.save(out_path)
    print(f"✅ Saved: {out_path}")
