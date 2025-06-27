import os
import time
from openpyxl import load_workbook
import winsound

# === Config ===
desktop_path = os.path.expanduser("~/Desktop")
file_path = os.path.join(desktop_path, "الخليل غياب تراكمي  26-06.xlsx")
sheet_name = "ادبي"
column = 'H'
start_row, end_row = 14, 109
expected_max = 85
check_interval = 5  # seconds

def get_non_empty_with_trigger_cell():
    wb = load_workbook(file_path, data_only=True)
    ws = wb[sheet_name]
    count = 0
    trigger_cell = None
    for row in range(start_row, end_row + 1):
        cell_ref = f"{column}{row}"
        value = ws[cell_ref].value
        if value is not None and str(value).strip() != "":
            count += 1
            if count > expected_max and trigger_cell is None:
                trigger_cell = cell_ref
    return count, trigger_cell

print(f"🔁 Monitoring {column}{start_row}:{column}{end_row}...")
print(f"🔔 Alert if non-empty values exceed {expected_max}\n")

try:
    while True:
        current_count, trigger_cell = get_non_empty_with_trigger_cell()
        if current_count > expected_max:
            print(f"⚠️ {current_count} entries detected! Limit = {expected_max}")
            print(f"🚨 Entry that triggered alert: {trigger_cell}")
            winsound.Beep(1000, 800)
        else:
            print(f"✅ OK: {current_count} entries")
        time.sleep(check_interval)
except KeyboardInterrupt:
    print("\n🛑 Monitoring stopped.")
