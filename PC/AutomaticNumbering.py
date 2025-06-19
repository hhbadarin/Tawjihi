from openpyxl import load_workbook
from openpyxl.styles import Border, Side, Alignment
import os

# Expand file paths
input_path = os.path.expanduser('~/Desktop/AA.xlsx')
output_path = os.path.expanduser('~/Desktop/updated_file.xlsx')

# Load workbook
wb = load_workbook(input_path)

# Define styles
thin_border = Border(bottom=Side(style='thin'))
center_wrap_align = Alignment(horizontal='center', vertical='center', wrap_text=True)

for sheet in wb.worksheets:
    ws = sheet

    # Set minimal margins (in inches)
    ws.page_margins.top = 0.85
    ws.page_margins.bottom = 0.6
    ws.page_margins.left = 0.4
    ws.page_margins.right = 0.4
    ws.page_margins.header = 0.35
    ws.page_margins.footer = 0.20

    # Center content horizontally on page when printing
    ws.page_setup.horizontalCentered = True

     # Fit all columns to 1 page wide, height can be multiple pages
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = False
    ws.page_setup.scale = 100

    # Set header/footer text
    #ws.oddHeader.left.text = "Your Left Header"
    #ws.oddHeader.right.text = "Your Right Header"
    ws.oddHeader.center.text = '&"Arial,Bold"&35 المكلفين بالعمل في الثانوية العامة للعام 2025'
    #ws.oddFooter.left.text = "Generated on &D"
    ws.oddFooter.right.text = "صفحة &P من &N"

    # Set header row height
    ws.row_dimensions[1].height = 45

    # Find "الرقم" column index
    header_row = 1
    column_index = None
    for col in range(1, ws.max_column + 1):
        if ws.cell(row=header_row, column=col).value == "الرقم":
            column_index = col
            break

    if column_index is not None:
        # Fill "الرقم" column and set row height + alignment
        for row in range(2, ws.max_row + 1):
            cell = ws.cell(row=row, column=column_index, value=row - 1)
            cell.alignment = center_wrap_align
            ws.row_dimensions[row].height = 40

    # Center-align all cells and add bottom border on last row
    last_row = ws.max_row
    for row in ws.iter_rows(min_row=1, max_row=last_row, max_col=ws.max_column):
        for cell in row:
            cell.alignment = center_wrap_align
            if cell.row == last_row:
                cell.border = Border(
                    left=cell.border.left,
                    right=cell.border.right,
                    top=cell.border.top,
                    bottom=thin_border.bottom
                )

    # Repeat first row on all printed pages
    ws.print_title_rows = '1:1'

# Save the modified workbook
wb.save(output_path)
