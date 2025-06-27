from openpyxl import load_workbook
from openpyxl.styles import Border, Side, Alignment, Font, PatternFill
from datetime import datetime
import os
import re

# Get today's date in the required formats
today = datetime.today()
today_str = today.strftime('%d-%m-%Y')             # e.g., 26-06-2025
today_short = today.strftime('%d-%m')              # e.g., 26-06

# Build dynamic file paths
input_path = os.path.expanduser(f'~/Desktop/hebron-{today_str}/{today_str} توزيع غياب الخليل.xlsx')
output_path = os.path.expanduser(f'~/Desktop/hebron-{today_str}/{today_short} توزيع غياب الخليل منسق.xlsx')

# Extract the district name from the input filename dynamically
match = re.search(r'توزيع غياب (.+)\.xlsx$', input_path)
district_name = match.group(1) if match else "المديرية"  # fallback if no match

# Load workbook
wb = load_workbook(input_path)

# Define styles
thin = Side(style='thin')
full_border = Border(top=thin, left=thin, right=thin, bottom=thin)
center_wrap_align = Alignment(horizontal='center', vertical='center', wrap_text=True)

header_font = Font(size=16, color='FFFFFF', bold=True)
cell_font = Font(size=13)
header_fill = PatternFill(fill_type="solid", start_color="4287f5")

for sheet in wb.worksheets:
    ws = sheet

    # Page setup
    ws.page_margins.top = 1.4
    ws.page_margins.bottom = 0.5
    ws.page_margins.left = 0.4
    ws.page_margins.right = 0.4
    ws.page_margins.header = 0.6  # increased to avoid overlap
    ws.page_margins.footer = 0.3
    ws.page_setup.paperSize = ws.PAPERSIZE_A4
    ws.page_setup.orientation = ws.ORIENTATION_PORTRAIT
    ws.page_setup.horizontalCentered = True
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 999
    ws.sheet_properties.pageSetUpPr.fitToPage = True

    # Header/footer with district and date on new line
    ws.oddHeader.center.text = f'&"Arial,Bold"&20 توزيع الغياب في مديرية {district_name}\n{today_str}'
    ws.oddFooter.right.text = "&12 صفحة &P من &N"
    ws.oddFooter.left.text = "&12 &D"

    # Header row height
    ws.row_dimensions[1].height = 45

    # Check/add "الرقم" column
    header_row = 1
    column_index = None
    for col in range(1, ws.max_column + 1):
        if ws.cell(row=header_row, column=col).value == "الرقم":
            column_index = col
            break

    if column_index is None:
        ws.insert_cols(1)
        ws.cell(row=1, column=1).value = "الرقم"
        column_index = 1

    # Fill numbering and apply alignment/row height
    for row in range(2, ws.max_row + 1):
        cell = ws.cell(row=row, column=column_index, value=row - 1)
        cell.alignment = center_wrap_align
        cell.font = cell_font
        ws.row_dimensions[row].height = 40

    # Apply alignment, font, borders, and header background fill
    for row in ws.iter_rows(min_row=1, max_row=ws.max_row, max_col=ws.max_column):
        for cell in row:
            cell.alignment = center_wrap_align
            cell.border = full_border
            if cell.row == 1:
                cell.font = header_font
                cell.fill = header_fill
            else:
                cell.font = cell_font

    # Repeat header row on print pages
    ws.print_title_rows = '1:1'

    # Auto-adjust column widths with max width limit
    max_col_width = 30
    for col in ws.columns:
        max_length = 0
        column_letter = col[0].column_letter
        for cell in col:
            try:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            except:
                pass
        adjusted_width = min(max_length + 4, max_col_width)
        ws.column_dimensions[column_letter].width = adjusted_width

# Save workbook
wb.save(output_path)
