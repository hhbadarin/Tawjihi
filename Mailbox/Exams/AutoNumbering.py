from openpyxl import load_workbook
import os

# Expand the user's home directory properly
input_path = os.path.expanduser('~/Desktop/AA.xlsx')
output_path = os.path.expanduser('~/Desktop/updated_file.xlsx')

# Load the Excel workbook
wb = load_workbook(input_path)

for sheet in wb.worksheets:
    ws = sheet

    # Find the column index for "الرقم"
    header_row = 1
    column_index = None
    for col in range(1, ws.max_column + 1):
        if ws.cell(row=header_row, column=col).value == "الرقم":
            column_index = col
            break

    if column_index is not None:
        # Fill the column with numbers from 1 to the last used row
        for row in range(2, ws.max_row + 1):  # Start from row 2 (after header)
            ws.cell(row=row, column=column_index, value=row - 1)

# Save the modified workbook
wb.save(output_path)
